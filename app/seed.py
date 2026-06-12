"""
Seed the database from the existing Excel file.
Usage: python -m app.seed
"""

from datetime import datetime, timedelta

import openpyxl
from dateutil import parser as dateparser

from app.auth import hash_password
from app.database import Base, SessionLocal, engine
from app.models import Participation, Race, RaceType, User

EXCEL_PATH = "TEST Calendario 2026 Valbelluna Motorsport.xlsx"

EXCEL_SERIAL_EPOCH = datetime(1899, 12, 30)


def excel_date_to_date(serial: float) -> datetime:
    return EXCEL_SERIAL_EPOCH + timedelta(days=int(serial))


def parse_date_cell(value) -> datetime | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        if value < 1:
            return None
        return excel_date_to_date(value)
    if isinstance(value, datetime):
        return value
    if isinstance(value, str):
        try:
            return dateparser.parse(value, dayfirst=True)
        except (ValueError, TypeError):
            return None
    return None


def seed():
    Base.metadata.create_all(bind=engine)
    db = SessionLocal()

    # Create default admin if missing
    admin = db.query(User).filter(User.ruolo == "admin").first()
    if not admin:
        admin = User(
            nome="Admin",
            email="admin@valbellunamotorsport.it",
            password_hash=hash_password("admin123"),
            ruolo="admin",
        )
        db.add(admin)
        db.commit()
        print("✅ Admin creato: admin@valbellunamotorsport.it / admin123")

    # Create SuperAdmin if missing
    superadmin = db.query(User).filter(User.ruolo == "superadmin").first()
    if not superadmin:
        superadmin = User(
            nome="SuperAdmin",
            email="superadmin@valbellunamotorsport.it",
            password_hash=hash_password("superadmin123"),
            ruolo="superadmin",
        )
        db.add(superadmin)
        db.commit()
        print(
            "✅ SuperAdmin creato: superadmin@valbellunamotorsport.it / superadmin123"
        )

    # Seed initial race types
    initial_types = [
        ("RA", "Raggruppamento A"),
        ("CR", "Campionato Regionale"),
        ("OR", "Open Race"),
        ("REG", "Regionare"),
    ]
    for codice, descrizione in initial_types:
        existing = db.query(RaceType).filter(RaceType.codice == codice).first()
        if not existing:
            db.add(RaceType(codice=codice, descrizione=descrizione))
    db.commit()

    # Read Excel
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    except FileNotFoundError:
        print(f"❌ File non trovato: {EXCEL_PATH}")
        return

    ws = wb.active
    if ws is None:
        print("❌ Nessun foglio attivo")
        return

    # Row 1: headers — extract member names from columns E..AC (indices 4..28)
    headers = [cell.value for cell in ws[1]]
    member_names = []
    member_cols = {}  # col_index -> name
    for i, h in enumerate(headers):
        if i >= 4 and i <= 28 and h:  # columns E through AC
            name = h.strip()
            member_names.append(name)
            member_cols[i] = name

    print(f"👥 Trovati {len(member_names)} membri")

    # Create members (skip if already exist)
    user_map = {}  # name -> User
    for name in member_names:
        existing = db.query(User).filter(User.nome == name).first()
        if existing:
            user_map[name] = existing
        else:
            email = name.lower().replace(" ", ".") + "@valbellunamotorsport.it"
            u = User(
                nome=name,
                email=email,
                password_hash=hash_password(name.lower().replace(" ", "")),
                ruolo="membro",
            )
            db.add(u)
            db.flush()
            user_map[name] = u
            print(f"  👤 Creato: {name} ({email})")

    db.commit()

    # Parse races from rows
    # Structure: each race has a "main" row with description, tipo_gara, stato, note_auto, scadenza_conferma
    # Followed by 0-2 sub-rows (additional days) with just date and participations
    races_created = 0
    i = 3  # start from row 3 (0-indexed: row 3 = 1-indexed row 3)
    max_row = ws.max_row or 1000

    while i <= max_row:
        row = [ws.cell(row=i, column=c).value for c in range(1, 43)]
        desc = row[2]  # Column C (descrizione)

        if not desc or str(desc).strip() == "":
            i += 1
            continue

        desc = str(desc).strip()

        # Skip summary and note rows
        if desc in ("TOTALE GARE", "") or desc.startswith("Attenzione"):
            i += 1
            continue

        tipo_gara = str(row[3]).strip() if row[3] else ""
        # Skip if tipo_gara is a long text (not a race type code)
        if len(tipo_gara) > 5:
            i += 1
            continue

        # Parse dates
        data_inizio = parse_date_cell(row[1])  # Column B
        data_fine = None

        # Stato
        stato_raw = str(row[32]).strip() if row[32] else ""  # Column AG
        if stato_raw in ("Confermato", "Annullato", "In attesa di conferma"):
            stato = stato_raw
        elif "annull" in stato_raw.lower():
            stato = "Annullato"
        else:
            stato = "In attesa di conferma"

        # Scadenza conferma
        scadenza = parse_date_cell(row[31])  # Column AF

        # Note auto
        note_auto = str(row[30]).strip() if row[30] else ""  # Column AE
        if note_auto and note_auto.startswith("Auto:"):
            note_auto = note_auto  # keep as is
        elif note_auto and len(note_auto) > 0:
            note_auto = f"Auto: {note_auto}"

        # Check if this race already exists (by description and year)
        existing_race = None
        if data_inizio:
            existing_race = (
                db.query(Race)
                .filter(
                    Race.descrizione == desc,
                    Race.data_inizio == data_inizio.date()
                    if hasattr(data_inizio, "date")
                    else Race.data_inizio == data_inizio,
                )
                .first()
            )

        if existing_race:
            race = existing_race
        else:
            race = Race(
                data_inizio=data_inizio
                if isinstance(data_inizio, datetime)
                else data_inizio,
                data_fine=data_fine,
                descrizione=desc,
                tipo_gara=tipo_gara if len(tipo_gara) <= 5 else "",
                scadenza_conferma=scadenza
                if isinstance(scadenza, datetime)
                else scadenza,
                stato=stato,
                note_auto=note_auto if note_auto else None,
            )
            db.add(race)
            db.flush()
            races_created += 1

        # Process participations for this row
        for col_idx, name in member_cols.items():
            val = row[col_idx]  # Column index in the row (0-based)
            if val and str(val).strip().lower() in ("s", "si", "x", "1"):
                status = "si"
            elif val and str(val).strip().lower() in ("n", "no", "0"):
                status = "no"
            else:
                status = "indeciso"

            user = user_map.get(name)
            if user:
                existing_p = (
                    db.query(Participation)
                    .filter(
                        Participation.user_id == user.id,
                        Participation.race_id == race.id,
                    )
                    .first()
                )
                if not existing_p:
                    db.add(
                        Participation(
                            user_id=user.id,
                            race_id=race.id,
                            status=status,
                        )
                    )

        i += 1

        # Check if next row is a sub-row (same race, additional day)
        if i <= max_row:
            next_row = [ws.cell(row=i, column=c).value for c in range(1, 43)]
            next_desc = next_row[2]
            if next_desc is None or str(next_desc).strip() == "":
                # This is a sub-row (same race, additional day)
                sub_date = parse_date_cell(next_row[1])
                if sub_date and data_inizio:
                    race.data_fine = (
                        sub_date if isinstance(sub_date, datetime) else sub_date
                    )
                    db.flush()

                for col_idx, name in member_cols.items():
                    val = next_row[col_idx]
                    if val and str(val).strip().lower() in ("s", "si", "x", "1"):
                        status = "si"
                    elif val and str(val).strip().lower() in ("n", "no", "0"):
                        status = "no"
                    else:
                        status = "indeciso"

                    user = user_map.get(name)
                    if user:
                        existing_p = (
                            db.query(Participation)
                            .filter(
                                Participation.user_id == user.id,
                                Participation.race_id == race.id,
                            )
                            .first()
                        )
                        if existing_p and status == "si" and existing_p.status != "si":
                            existing_p.status = "si"
                        elif not existing_p:
                            db.add(
                                Participation(
                                    user_id=user.id,
                                    race_id=race.id,
                                    status=status,
                                )
                            )

                i += 1

        db.commit()

    db.close()
    print(f"\n✅ Seed completato! {races_created} gare importate.")


if __name__ == "__main__":
    seed()
